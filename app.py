# app.py — Metal Débit Web (version “niveau .exe”)
# Dépendances: streamlit, openpyxl, reportlab
# Recommandé sur Streamlit Cloud: ajoute runtime.txt => python-3.11

from __future__ import annotations

import io
import random
from dataclasses import dataclass
from datetime import date
from typing import Dict, List, Tuple

import streamlit as st

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


# =========================
# CONFIG / AUTH
# =========================
st.set_page_config(page_title="Metal Débit", page_icon="🧰", layout="wide")

# ⚠️ Change ton mot de passe ici (ou via st.secrets)
PASSWORD = "metaldebit123"


def gate_password() -> None:
    if "auth_ok" not in st.session_state:
        st.session_state.auth_ok = False

    if st.session_state.auth_ok:
        return

    st.title("🔒 Accès sécurisé — Metal Débit")
    st.caption("Entrez le mot de passe pour accéder à l’application.")
    pwd = st.text_input("Mot de passe", type="password")

    c1, c2 = st.columns([1, 2])
    with c1:
        if st.button("Connexion", use_container_width=True):
            if pwd == PASSWORD:
                st.session_state.auth_ok = True
                st.rerun()
            else:
                st.error("Mot de passe incorrect")
    with c2:
        st.info("Tu peux changer le mot de passe dans app.py (PASSWORD = ...).")

    st.stop()


gate_password()


# =========================
# MODELES
# =========================
@dataclass
class Profile:
    name: str
    bar_length_mm: int = 6000
    price_ht_per_bar: float = 0.0


@dataclass(frozen=True)
class Piece:
    length: int


@dataclass
class Bar:
    pieces: List[Piece]
    used: int
    leftover: int
    source: str          # "bar" (barre neuve) ou "stock" (chute utilisée)
    source_length: int   # longueur source


# =========================
# OPTIM
# =========================
def expand_pieces(requirements: Dict[int, int]) -> List[Piece]:
    out: List[Piece] = []
    for length, qty in requirements.items():
        out.extend([Piece(int(length))] * int(qty))
    return out


def pack_first_fit(pieces: List[Piece], bar_length: int, kerf: int) -> List[Bar]:
    bars: List[List[Piece]] = []
    used: List[int] = []

    for p in pieces:
        placed = False
        for i in range(len(bars)):
            add = p.length + (kerf if len(bars[i]) > 0 else 0)
            if used[i] + add <= bar_length:
                bars[i].append(p)
                used[i] += add
                placed = True
                break
        if not placed:
            if p.length > bar_length:
                raise ValueError(f"Pièce {p.length} mm > barre {bar_length} mm (impossible).")
            bars.append([p])
            used.append(p.length)

    result: List[Bar] = []
    for ps, u in zip(bars, used):
        result.append(Bar(ps, u, bar_length - u, source="bar", source_length=bar_length))
    return result


def pack_into_stock_then_bars(
    pieces: List[Piece],
    stock: List[int],
    bar_length: int,
    kerf: int
) -> Tuple[List[Bar], List[int]]:
    stock_sorted = sorted([int(x) for x in stock if int(x) > 0], reverse=True)

    stock_bins: List[List[Piece]] = [[] for _ in stock_sorted]
    stock_used: List[int] = [0 for _ in stock_sorted]
    remaining: List[Piece] = []

    # best-fit stock (pièces décroissantes)
    for p in sorted(pieces, key=lambda x: x.length, reverse=True):
        best_idx = None
        best_rem = None
        for i, L in enumerate(stock_sorted):
            add = p.length + (kerf if len(stock_bins[i]) > 0 else 0)
            if stock_used[i] + add <= L:
                rem = L - (stock_used[i] + add)
                if best_rem is None or rem < best_rem:
                    best_rem = rem
                    best_idx = i

        if best_idx is None:
            remaining.append(p)
        else:
            add = p.length + (kerf if len(stock_bins[best_idx]) > 0 else 0)
            stock_bins[best_idx].append(p)
            stock_used[best_idx] += add

    bars: List[Bar] = []
    new_stock: List[int] = []

    for i, L in enumerate(stock_sorted):
        if len(stock_bins[i]) == 0:
            new_stock.append(L)
        else:
            leftover = L - stock_used[i]
            bars.append(Bar(stock_bins[i], stock_used[i], leftover, source="stock", source_length=L))
            if leftover > 0:
                new_stock.append(leftover)

    bars.extend(pack_first_fit(remaining, bar_length, kerf))

    bars = sorted(bars, key=lambda b: b.leftover)
    new_stock = sorted([x for x in new_stock if x > 0], reverse=True)
    return bars, new_stock


def score_solution(
    bars: List[Bar],
    scrap_max: int,
    reusable_min: int,
    strategy: str
) -> float:
    nb_new_bars = sum(1 for b in bars if b.source == "bar")
    leftovers = [b.leftover for b in bars]
    total_waste = sum(leftovers)
    mid = [w for w in leftovers if scrap_max < w < reusable_min]
    mid_penalty = sum(mid)
    max_leftover = max(leftovers) if leftovers else 0

    if strategy == "min_barres":
        return nb_new_bars * 1e12 + total_waste * 1e3 + mid_penalty * 10 - max_leftover

    if strategy == "max_grosse_chute":
        # favorise une grosse chute réutilisable (ex: 4310) plutôt que 470/540
        return nb_new_bars * 1e12 + total_waste * 1e3 + mid_penalty * 70 - (max_leftover * 1e5)

    if strategy == "zero_micro":
        # punit fortement les chutes "moyennes"
        return nb_new_bars * 1e12 + total_waste * 1e3 + mid_penalty * 250 - max_leftover

    # equilibre
    return nb_new_bars * 1e12 + total_waste * 1e3 + mid_penalty * 90 - max_leftover


def optimize_profile(
    requirements: Dict[int, int],
    stock: List[int],
    bar_length: int,
    kerf: int,
    iterations: int,
    scrap_max: int,
    reusable_min: int,
    strategy: str,
    seed: int = 42
) -> Tuple[List[Bar], List[int]]:
    pieces = expand_pieces(requirements)
    current = sorted(pieces, key=lambda p: p.length, reverse=True)

    best_bars, best_stock = pack_into_stock_then_bars(current, stock, bar_length, kerf)
    best_score = score_solution(best_bars, scrap_max, reusable_min, strategy)

    rng = random.Random(seed)

    # Itérations = nombre d'essais (échanges aléatoires)
    for _ in range(max(1, iterations)):
        if len(current) < 2:
            break

        new_order = current[:]
        i = rng.randrange(len(new_order))
        j = rng.randrange(len(new_order))
        new_order[i], new_order[j] = new_order[j], new_order[i]

        bars, new_stock = pack_into_stock_then_bars(new_order, stock, bar_length, kerf)
        sc = score_solution(bars, scrap_max, reusable_min, strategy)
        if sc < best_score:
            best_score = sc
            best_bars, best_stock = bars, new_stock
            current = new_order

    best_bars = sorted(best_bars, key=lambda b: b.leftover)
    best_stock = sorted([int(x) for x in best_stock if int(x) > 0], reverse=True)
    return best_bars, best_stock


# =========================
# HELPERS
# =========================
def bars_to_counts(bar: Bar) -> Dict[int, int]:
    counts: Dict[int, int] = {}
    for p in bar.pieces:
        counts[p.length] = counts.get(p.length, 0) + 1
    return counts


def parse_stock_csv(s: str) -> List[int]:
    if not s or not s.strip():
        return []
    out: List[int] = []
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            v = int(part)
            if v > 0:
                out.append(v)
        except Exception:
            continue
    return sorted(out, reverse=True)


def req_from_rows(rows: List[dict]) -> Dict[int, int]:
    req: Dict[int, int] = {}
    for r in rows:
        L = r.get("Longueur (mm)")
        Q = r.get("Quantité")

        if (L is None or L == "") and (Q is None or Q == ""):
            continue

        try:
            Li = int(L)
            Qi = int(Q)
            if Li > 0 and Qi > 0:
                req[Li] = req.get(Li, 0) + Qi
        except Exception:
            continue
    return req


# =========================
# EXPORTS
# =========================
def export_excel_like_table(
    project_name: str,
    client: str,
    created: str,
    kerf: int,
    strategy: str,
    profiles: Dict[str, Profile],
    results: Dict[str, List[Bar]],
    max_coupe_cols: int = 7
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Feuil1"

    dark_fill = PatternFill("solid", fgColor="404040")
    white_font_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="777777")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["B2"] = "Projet"; ws["C2"] = project_name
    ws["B3"] = "Client"; ws["C3"] = client
    ws["B4"] = "Date"; ws["C4"] = created
    ws["B5"] = "Stratégie"; ws["C5"] = strategy
    ws["B6"] = "Trait de scie"; ws["C6"] = f"{kerf} mm"
    for cell in ["B2", "B3", "B4", "B5", "B6"]:
        ws[cell].font = bold

    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 34

    main_row = 8
    main_col = 2  # B

    headers = ["barre index", "profil"] + ["coupe"] * max_coupe_cols + ["chutes/mm"]
    for i, h in enumerate(headers):
        c = ws.cell(row=main_row, column=main_col + i, value=h)
        c.fill = dark_fill
        c.font = white_font_bold
        c.alignment = center
        c.border = border

    widths = [14, 14] + [12] * max_coupe_cols + [12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(main_col + i)].width = w

    r = main_row + 1
    global_index = 1

    recap: Dict[str, Dict[str, float]] = {}
    total_ht = 0.0

    for prof_name, bars in results.items():
        prof = profiles[prof_name]
        nb_new = sum(1 for b in bars if b.source == "bar")
        total_ht += nb_new * prof.price_ht_per_bar
        recap[prof_name] = {"nb_barres": nb_new, "prix_ht": prof.price_ht_per_bar}

        for bar in bars:
            ws.cell(row=r, column=main_col, value=f"barre {global_index}").alignment = left
            ws.cell(row=r, column=main_col + 1, value=prof_name).alignment = left

            counts = bars_to_counts(bar)
            coupes = [f"{L}*{q}" for L, q in sorted(counts.items(), reverse=True)]
            for j in range(max_coupe_cols):
                ws.cell(
                    row=r,
                    column=main_col + 2 + j,
                    value=(coupes[j] if j < len(coupes) else "")
                ).alignment = center

            ws.cell(row=r, column=main_col + 2 + max_coupe_cols, value=bar.leftover).alignment = center

            for ccol in range(main_col, main_col + len(headers)):
                ws.cell(row=r, column=ccol).border = border

            r += 1
            global_index += 1

    resume_col = main_col + len(headers) + 2
    resume_row = main_row

    resume_headers = ["profils", "nb de barres", "prix/ht"]
    for i, h in enumerate(resume_headers):
        c = ws.cell(row=resume_row, column=resume_col + i, value=h)
        c.fill = dark_fill
        c.font = white_font_bold
        c.alignment = center
        c.border = border

    ws.column_dimensions[get_column_letter(resume_col)].width = 14
    ws.column_dimensions[get_column_letter(resume_col + 1)].width = 12
    ws.column_dimensions[get_column_letter(resume_col + 2)].width = 12

    rr = resume_row + 1
    for prof_name, info in recap.items():
        ws.cell(row=rr, column=resume_col, value=prof_name).alignment = left
        ws.cell(row=rr, column=resume_col + 1, value=int(info["nb_barres"])).alignment = center
        cprice = ws.cell(row=rr, column=resume_col + 2, value=float(info["prix_ht"]))
        cprice.number_format = '#,##0.00" €"'
        cprice.alignment = center

        for ccol in range(resume_col, resume_col + 3):
            ws.cell(row=rr, column=ccol).border = border

        rr += 1

    label_row = rr + 1
    label_cell = ws.cell(row=label_row, column=resume_col + 1, value="prix total ht")
    label_cell.fill = dark_fill
    label_cell.font = white_font_bold
    label_cell.alignment = center
    label_cell.border = border

    value_cell = ws.cell(row=label_row + 1, column=resume_col + 1, value=total_ht)
    value_cell.number_format = '#,##0.00" €"'
    value_cell.alignment = center
    value_cell.border = border

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf_atelier_no_price(
    project_name: str,
    client: str,
    created: str,
    kerf: int,
    strategy: str,
    profiles: Dict[str, Profile],
    results: Dict[str, List[Bar]]
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm
    )

    styles = getSampleStyleSheet()
    s_title = styles["Title"]
    s_norm = styles["Normal"]
    story = []

    story.append(Paragraph("Bon de débit – métal", s_title))
    story.append(Spacer(1, 4 * mm))

    info = (
        f"<b>Projet :</b> {project_name} &nbsp;&nbsp; "
        f"<b>Client :</b> {client}<br/>"
        f"<b>Date :</b> {created} &nbsp;&nbsp; "
        f"<b>Trait de scie :</b> {kerf} mm &nbsp;&nbsp; "
        f"<b>Stratégie :</b> {strategy}"
    )
    story.append(Paragraph(info, s_norm))
    story.append(Spacer(1, 6 * mm))

    def bar_desc(b: Bar) -> str:
        counts = bars_to_counts(b)
        items = [f"{L}×{q}" for L, q in sorted(counts.items(), reverse=True)]
        return " + ".join(items)

    for prof_name, bars in results.items():
        prof = profiles.get(prof_name)

        story.append(Paragraph(f"<b>Profil :</b> {prof_name}", styles["Heading2"]))
        if prof:
            story.append(Paragraph(f"Longueur barre neuve : {prof.bar_length_mm} mm", s_norm))
        story.append(Spacer(1, 3 * mm))

        data = [["OK", "Barre", "Source", "Longueur", "Découpes", "Chute (mm)"]]
        for i, b in enumerate(bars, 1):
            src = "STOCK" if b.source == "stock" else "BARRE"
            data.append(["☐", f"{i:02d}", src, str(b.source_length), bar_desc(b), str(b.leftover)])

        tbl = Table(data, colWidths=[10 * mm, 14 * mm, 18 * mm, 22 * mm, 110 * mm, 20 * mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#404040")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),

            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#888888")),
            ("FONTSIZE", (0, 1), (-1, -1), 9),

            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("ALIGN", (1, 1), (3, -1), "CENTER"),
            ("ALIGN", (5, 1), (5, -1), "CENTER"),
            ("ALIGN", (4, 1), (4, -1), "LEFT"),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 8 * mm))

    doc.build(story)
    return buf.getvalue()


# =========================
# STATE INIT
# =========================
def init_state() -> None:
    if "project_name" not in st.session_state:
        st.session_state.project_name = "Projet"
    if "client_name" not in st.session_state:
        st.session_state.client_name = "Client"
    if "project_date" not in st.session_state:
        st.session_state.project_date = str(date.today())

    if "kerf_mm" not in st.session_state:
        st.session_state.kerf_mm = 0
    if "iterations" not in st.session_state:
        st.session_state.iterations = 30000
    if "scrap_max_mm" not in st.session_state:
        st.session_state.scrap_max_mm = 100
    if "reusable_min_mm" not in st.session_state:
        st.session_state.reusable_min_mm = 1500
    if "strategy" not in st.session_state:
        st.session_state.strategy = "equilibre"

    if "profiles" not in st.session_state:
        st.session_state.profiles = {}  # name -> Profile
    if "profile_rows" not in st.session_state:
        st.session_state.profile_rows = {}  # name -> list[dict] (10 lignes)
    if "profile_stock" not in st.session_state:
        st.session_state.profile_stock = {}  # name -> "4310,2500"

    if "results" not in st.session_state:
        st.session_state.results = {}  # name -> list[Bar]
    if "stock_after" not in st.session_state:
        st.session_state.stock_after = {}  # name -> list[int]


init_state()


# =========================
# UI — HEADER
# =========================
st.title("🧰 Metal Débit — Web (complet)")
st.caption("Multi-profils • 10 lignes ergonomiques • Stock chutes • Stratégies • Excel (prix) + PDF atelier (sans prix)")

h1, h2, h3, h4 = st.columns([2, 2, 1.2, 1.2])
with h1:
    st.session_state.project_name = st.text_input("Nom du projet", value=st.session_state.project_name, key="proj_name")
with h2:
    st.session_state.client_name = st.text_input("Client", value=st.session_state.client_name, key="client_name")
with h3:
    st.session_state.project_date = st.text_input("Date", value=st.session_state.project_date, key="proj_date")
with h4:
    if st.button("🧹 Reset résultats", use_container_width=True, key="reset_results"):
        st.session_state.results = {}
        st.session_state.stock_after = {}
        st.success("Résultats effacés.")

st.divider()

left, right = st.columns([1.05, 1.55], gap="large")

# =========================
# UI — PARAMS + ADD PROFILE
# =========================
with left:
    st.subheader("Paramètres atelier (globaux)")

    st.session_state.kerf_mm = st.number_input("Trait de scie (mm)", min_value=0, value=int(st.session_state.kerf_mm), step=1, key="kerf")
    st.session_state.iterations = st.number_input(
        "Itérations",
        min_value=0,
        value=int(st.session_state.iterations),
        step=1000,
        key="iters",
        help="Plus c'est haut, plus l'optimisation cherche une meilleure combinaison (mais c'est plus lent)."
    )
    st.session_state.scrap_max_mm = st.number_input("Chute poubelle ≤ (mm)", min_value=0, value=int(st.session_state.scrap_max_mm), step=10, key="scrapmax")
    st.session_state.reusable_min_mm = st.number_input("Chute réutilisable ≥ (mm)", min_value=0, value=int(st.session_state.reusable_min_mm), step=50, key="reusemin")

    st.session_state.strategy = st.selectbox(
        "Stratégie",
        ["equilibre", "min_barres", "max_grosse_chute", "zero_micro"],
        index=["equilibre", "min_barres", "max_grosse_chute", "zero_micro"].index(st.session_state.strategy),
        key="strategy",
        help=(
            "equilibre: bon compromis\n"
            "min_barres: minimise d'abord le nombre de barres\n"
            "max_grosse_chute: favorise une grosse chute réutilisable\n"
            "zero_micro: punit très fort les chutes “moyennes”"
        )
    )

    if st.session_state.reusable_min_mm <= st.session_state.scrap_max_mm:
        st.warning("Conseil: mets 'réutilisable' > 'poubelle' (ex: poubelle=100, réutilisable=1500).")

    st.divider()
    st.subheader("Ajouter / modifier un profil")

    p_name = st.text_input("Nom profil (ex: 30x30x2)", value="", key="p_name")
    p_bar = st.number_input("Longueur barre neuve (mm)", min_value=1, value=6000, step=10, key="p_bar")
    p_price = st.number_input("Prix HT / barre (Excel)", min_value=0.0, value=0.0, step=1.0, key="p_price")

    b1, b2 = st.columns(2)
    with b1:
        if st.button("➕ Ajouter / Mettre à jour", use_container_width=True, key="add_update_profile"):
            name = p_name.strip()
            if not name:
                st.error("Nom profil obligatoire.")
            else:
                st.session_state.profiles[name] = Profile(name=name, bar_length_mm=int(p_bar), price_ht_per_bar=float(p_price))
                if name not in st.session_state.profile_rows:
                    st.session_state.profile_rows[name] = [{"Longueur (mm)": "", "Quantité": ""} for _ in range(10)]
                if name not in st.session_state.profile_stock:
                    st.session_state.profile_stock[name] = ""
                st.success(f"Profil '{name}' enregistré.")
    with b2:
        if st.button("🗑️ Supprimer", use_container_width=True, key="delete_profile"):
            name = p_name.strip()
            if name in st.session_state.profiles:
                st.session_state.profiles.pop(name, None)
                st.session_state.profile_rows.pop(name, None)
                st.session_state.profile_stock.pop(name, None)
                st.success(f"Profil '{name}' supprimé.")
            else:
                st.info("Entre le nom exact du profil à supprimer (champ 'Nom profil').")

# =========================
# UI — PROFILES INPUTS
# =========================
with right:
    st.subheader("Profils & pièces (10 lignes par profil)")

    if not st.session_state.profiles:
        st.info("Ajoute un profil à gauche pour commencer.")
    else:
        for prof_name, prof in st.session_state.profiles.items():
            with st.expander(
                f"Profil: {prof_name}  |  Barre: {prof.bar_length_mm} mm  |  Prix HT/barre: {prof.price_ht_per_bar:.2f} €",
                expanded=False
            ):
                st.caption("Remplis seulement les lignes utiles. Les lignes vides sont ignorées.")

                if prof_name not in st.session_state.profile_rows or len(st.session_state.profile_rows[prof_name]) != 10:
                    st.session_state.profile_rows[prof_name] = [{"Longueur (mm)": "", "Quantité": ""} for _ in range(10)]

                edited = st.data_editor(
                    st.session_state.profile_rows[prof_name],
                    num_rows="fixed",
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Longueur (mm)": st.column_config.NumberColumn("Longueur (mm)", min_value=0, step=1),
                        "Quantité": st.column_config.NumberColumn("Quantité", min_value=0, step=1),
                    },
                    key=f"editor_{prof_name}"
                )
                st.session_state.profile_rows[prof_name] = edited

                st.write("Stock de chutes réutilisables (mm), séparées par virgules (ex: 4310, 2500, 1800)")
                st.session_state.profile_stock[prof_name] = st.text_input(
                    "Chutes en stock",
                    value=st.session_state.profile_stock.get(prof_name, ""),
                    key=f"stock_{prof_name}"
                )

st.divider()

# =========================
# ACTIONS — CALCUL + EXPORT
# =========================
a1, a2, a3, a4 = st.columns([1.2, 1.2, 1.2, 1.2])

with a1:
    do_calc = st.button("🧮 Calculer tout", use_container_width=True, key="calc_all")
with a2:
    show_details = st.checkbox("Afficher détails des barres", value=True, key="show_details")
with a3:
    show_stock_after = st.checkbox("Afficher stock après calcul", value=True, key="show_stock_after")
with a4:
    st.caption("Excel = avec prix • PDF atelier = sans prix")

if do_calc:
    if not st.session_state.profiles:
        st.error("Ajoute au moins un profil.")
    else:
        results: Dict[str, List[Bar]] = {}
        stock_after: Dict[str, List[int]] = {}
        errors: List[str] = []

        for prof_name, prof in st.session_state.profiles.items():
            rows = st.session_state.profile_rows.get(prof_name, [{"Longueur (mm)": "", "Quantité": ""} for _ in range(10)])
            req = req_from_rows(rows)
            if not req:
                continue

            stock = parse_stock_csv(st.session_state.profile_stock.get(prof_name, ""))
            try:
                bars, new_stock = optimize_profile(
                    requirements=req,
                    stock=stock,
                    bar_length=prof.bar_length_mm,
                    kerf=int(st.session_state.kerf_mm),
                    iterations=int(st.session_state.iterations),
                    scrap_max=int(st.session_state.scrap_max_mm),
                    reusable_min=int(st.session_state.reusable_min_mm),
                    strategy=str(st.session_state.strategy),
                )
                results[prof_name] = bars
                stock_after[prof_name] = new_stock
            except Exception as e:
                errors.append(f"{prof_name}: {e}")

        if errors:
            st.error("Erreurs:\n" + "\n".join(errors))

        st.session_state.results = results
        st.session_state.stock_after = stock_after

        if not results:
            st.warning("Aucun calcul: pas de pièces saisies (ou tout vide).")
        else:
            st.success("Calcul terminé.")

# =========================
# RENDER RESULTS
# =========================
results = st.session_state.results

if results:
    st.subheader("Résultats")

    total_new_bars = 0
    total_ht = 0.0

    for prof_name, bars in results.items():
        prof = st.session_state.profiles[prof_name]
        nb_new = sum(1 for b in bars if b.source == "bar")
        waste = sum(b.leftover for b in bars)
        cost = nb_new * prof.price_ht_per_bar

        total_new_bars += nb_new
        total_ht += cost

        st.markdown(f"### {prof_name}")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Barres neuves", nb_new)
        c2.metric("Chute totale (mm)", waste)
        c3.metric("Longueur barre", f"{prof.bar_length_mm} mm")
        c4.metric("Coût HT (Excel)", f"{cost:.2f} €")

        if show_stock_after:
            new_stock = st.session_state.stock_after.get(prof_name, [])
            st.write("Stock chutes après calcul:", new_stock)

        if show_details:
            lines = []
            for i, bar in enumerate(bars, 1):
                counts = bars_to_counts(bar)
                desc = " + ".join(f"{L}x{q}" for L, q in sorted(counts.items(), reverse=True))
                src = "STOCK" if bar.source == "stock" else "BARRE"
                lines.append(f"{src} {bar.source_length}mm | Barre {i:02d}: {desc} | chute {bar.leftover}mm")
            st.code("\n".join(lines), language="text")

    st.divider()
    t1, t2, t3 = st.columns(3)
    t1.metric("TOTAL barres neuves", total_new_bars)
    t2.metric("TOTAL HT (Excel)", f"{total_ht:.2f} €")
    t3.metric("Stratégie", st.session_state.strategy)

    # =========================
    # EXPORT BUTTONS
    # =========================
    st.subheader("Exports")

    excel_bytes = export_excel_like_table(
        project_name=st.session_state.project_name,
        client=st.session_state.client_name,
        created=st.session_state.project_date,
        kerf=int(st.session_state.kerf_mm),
        strategy=str(st.session_state.strategy),
        profiles=st.session_state.profiles,
        results=results,
        max_coupe_cols=7
    )

    pdf_bytes = export_pdf_atelier_no_price(
        project_name=st.session_state.project_name,
        client=st.session_state.client_name,
        created=st.session_state.project_date,
        kerf=int(st.session_state.kerf_mm),
        strategy=str(st.session_state.strategy),
        profiles=st.session_state.profiles,
        results=results
    )

    file_base = f"{st.session_state.project_name}_{st.session_state.client_name}_{st.session_state.project_date}".replace(" ", "_")

    e1, e2 = st.columns(2)
    with e1:
        st.download_button(
            "📥 Télécharger Excel (avec prix)",
            data=excel_bytes,
            file_name=f"{file_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with e2:
        st.download_button(
            "📥 Télécharger PDF atelier (sans prix)",
            data=pdf_bytes,
            file_name=f"{file_base}.pdf",
            mime="application/pdf",
            use_container_width=True
        )

else:
    st.info("Renseigne tes profils et pièces, puis clique **Calculer tout**.")
